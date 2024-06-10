import os
from kivy import Config

Config.set('graphics', 'multisamples', '0')
os.environ['KIVY_GL_BACKEND'] = 'angle_sdl2'

from kivy.uix.filechooser import FileChooser
import time
from kivy.properties import ListProperty, NumericProperty, StringProperty, Clock
from kivy.uix.anchorlayout import AnchorLayout
import kivy
from kivymd.app import MDApp
from kivy.core.window import Window
from kivy.lang.builder import Builder
from kivy.uix.screenmanager import ScreenManager, Screen
from plyer import filechooser
import tensorflow as tf
import numpy as np
import cv2
import pandas as pd
import datetime
from openpyxl import Workbook, load_workbook
from kivy.clock import Clock

Window.size = (360, 650)
# Builder.load_file("TBDetection.kv")


prediction = ""

# this holds the image to be verified
picture = ""

# loads the pre-trained model
loaded_model = tf.keras.models.load_model(
    r'C:\Users\victor.CHRISTTECH\Downloads\ProjectFolder\ImageClassification-main\models\my_tb_model.h5')


class HomeScreen(Screen):

    def file_chooser(self):
        filechooser.open_file(on_selection=self.selected,
                              filters=["*.png", "*.jpg", "*.bmp"], )

    def selected(self, selection):
        # this assigns the image_path to the picture variable
        global picture
        if selection:
            picture = selection[0]

    # changes the icon of the toggle button
    def change_toggle_button(self):
        button = self.ids.switch_button_icon
        button.icon = 'toggle-switch-off-outline' if button.icon == 'toggle-switch' else 'toggle-switch'

    def on_button_press(self):
        load_multiple_image = self.manager.get_screen('Multiple')
        load_multiple_image.load_multiple_files()

    def reload(self):
        self.manager.current = "Progress"
        #time.sleep(1)
        self.manager.current = "Home"

    def splash(self, *args):
        ScreenManager.current = "Home"


class ProgressScreen(Screen):
    pass

class UploadScreen(Screen):

    def change_image(self):
        global picture
        self.ids.display_image.source = f"{picture}"

    def verify_image(self):
        try:
            global picture, prediction
            img = cv2.imread(f"{picture}")
            resize = tf.image.resize(img, (256, 256))
            input_img = np.expand_dims(resize / 255, 0)

            prediction = loaded_model.predict(input_img)

            if prediction > 0.5:
                self.manager.current = "Detected"
                time.sleep(1.5)
                self.manager.transition.direction = "left"

            else:
                self.manager.current = "NotDetected"
                time.sleep(1.5)
                self.manager.transition.direction = "left"
        except:
            self.manager.current = 'Upload'

    def clear_fields(self):
        self.ids.patient_description.text = ""
        self.ids.patient_name.text = ""
        self.ids.display_image.source = ""

class TBDetectedScreen(Screen):

    def on_enter(self):
        global prediction
        print(prediction)
        # noinspection PyTypeChecker
        self.ids.prediction_accuracy.text = f"Patient has {round(prediction[0][0] * 100, 2)}% chance of TB"

    def save_patient_record(self):
        global prediction
        print(prediction)

        try:
            if prediction > 0.5:
                case = f"Patient has {round(prediction[0][0] * 100, 2)}% chance of TB"

            else:
                case = f"patient has {round(prediction[0][0] * 100, 2)}% chance of TB"

            """this section adds the patient information to an excel sheet"""
            excel_file_name = "diagnosesses.xlsx"
            excel_path = os.getcwd() + "\\" + excel_file_name

            if not os.path.exists(excel_path):
                workbook = Workbook()
                sheet = workbook.active

                sheet["A1"] = f"{self.manager.get_screen("Upload").ids.patient_name.text}"
                sheet["B1"] = f"{self.manager.get_screen("Upload").ids.patient_description.text}"
                sheet["C1"] = f"{case}"

                filename = excel_file_name
                workbook.save(filename=filename)

            else:
                work_book = excel_file_name
                file = load_workbook(work_book)
                page = file.active

                new_patient_data = [self.manager.get_screen("Upload").ids.patient_name.text,
                                    self.manager.get_screen("Upload").ids.patient_description.text, case]

                row = page.max_row + 1
                column = 1
                for i, info in enumerate(new_patient_data, start=row):
                    page.cell(row=row, column=column, value=info)

                    column += 1

                file.save(filename=work_book)

            self.manager.current = "Home"
            self.manager.transition.direction = "right"

        # this is triggered when the patient information can't be saved to the excelsheet
        except PermissionError:
            self.ids.prediction_accuracy.text = "Can't save: close excel sheet and try again..."
            self.ids.prediction_accuracy.color = (1, 0, 0, 1)
            if self.manager.current == "Detected":
                self.manager.current = "Detected"
            else:
                self.manager.current = "NotDetected"


class MultipleImagesScreen(Screen):

    def load_multiple_files(self):
        """"A Function that allows the user to select multiple image files"""

        filechooser.open_file(
            multiple=True,
            filters=["image files", "*.png", "*.jpg", "*.bmp"],
            on_selection=self.on_file_selection
        )

    def on_file_selection(self, selection):
        """A callback function that is called when a file is selected."""
        global prediction
        max_files = 10  # Set your desired maximum number of files
        selected_files = len(selection)

        if selected_files > max_files:
            use_files = selection[:max_files]
        else:
            use_files = selection
        self.selected_files = use_files

        def multiple_image_verify():

            try:
                count = 1

                for files in self.selected_files:
                    img = cv2.imread(f"{files}")

                    if img is None:
                        print(f"Error: Unable to load image {files}. Check file path and file integrity.")
                        continue

                    resize = tf.image.resize(img, (256, 256))
                    input_img = np.expand_dims(resize / 255, 0)

                    prediction = loaded_model.predict(input_img)

                    if prediction > 0.5:
                        print("image predicted to have TB")
                        print(files)

                        if count == 1:
                            self.ids.space_1.shadow_color = "red"
                            self.ids.display_image_1.source = f"{files}"

                        elif count == 2:
                            self.ids.space_2.shadow_color = "red"
                            self.ids.display_image_2.source = f"{files}"
                        elif count == 3:
                            self.ids.space_3.shadow_color = "red"
                            self.ids.display_image_3.source = f"{files}"
                        elif count == 4:
                            self.ids.space_4.shadow_color = "red"
                            self.ids.display_image_4.source = f"{files}"
                        elif count == 5:
                            self.ids.space_5.shadow_color = "red"
                            self.ids.display_image_5.source = f"{files}"
                        elif count == 6:
                            self.ids.space_6.shadow_color = "red"
                            self.ids.display_image_6.source = f"{files}"
                        elif count == 7:
                            self.ids.space_7.shadow_color = "red"
                            self.ids.display_image_7.source = f"{files}"
                        elif count == 8:
                            self.ids.space_8.shadow_color = "red"
                            self.ids.display_image_8.source = f"{files}"
                        elif count == 9:
                            self.ids.space_9.shadow_color = "red"
                            self.ids.display_image_9.source = f"{files}"
                        else:
                            self.ids.space_10.shadow_color = "red"
                            self.ids.display_image_10.source = f"{files}"
                        count += 1

                    else:
                        print(f"image predicted to not have TB")
                        if count == 1:
                            self.ids.space_1.shadow_color = "green"
                            self.ids.display_image_1.source = f"{files}"
                        elif count == 2:
                            self.ids.space_2.shadow_color = "green"
                            self.ids.display_image_2.source = f"{files}"
                        elif count == 3:
                            self.ids.space_3.shadow_color = "green"
                            self.ids.display_image_3.source = f"{files}"
                        elif count == 4:
                            self.ids.space_4.shadow_color = "green"
                            self.ids.display_image_4.source = f"{files}"
                        elif count == 5:
                            self.ids.space_5.shadow_color = "green"
                            self.ids.display_image_5.source = f"{files}"
                        elif count == 6:
                            self.ids.space_6.shadow_color = "green"
                            self.ids.display_image_6.source = f"{files}"
                        elif count == 7:
                            self.ids.space_7.shadow_color = "green"
                            self.ids.display_image_7.source = f"{files}"
                        elif count == 8:
                            self.ids.space_8.shadow_color = "green"
                            self.ids.display_image_8.source = f"{files}"
                        elif count == 9:
                            self.ids.space_9.shadow_color = "green"
                            self.ids.display_image_9.source = f"{files}"
                        else:
                            self.ids.space_10.shadow_color = "green"
                            self.ids.display_image_10.source = f"{files}"
                        count += 1

                        continue

            except Exception as e:
                print(f"there was an error: {e}")

        multiple_image_verify()


class TBNotDetectedScreen(Screen):
    def on_enter(self):
        global prediction
        self.ids.prediction_accuracy.text = f"Patient has {round(prediction[0][0] * 100, 2)}% chance of TB"

    def save_patient_record(self):
        TBDetectedScreen.save_patient_record(self)


class ProgressBarScreen(Screen):
    pass


class CircularProgressBar(AnchorLayout):
    set_value = NumericProperty(0)
    value = NumericProperty(0)
    bar_color = ListProperty([20 / 255, 115 / 255, 233 / 255])
    bar_width = NumericProperty(10)
    text = StringProperty("0%")
    duration = NumericProperty(1.5)
    counter = 0

    def __init__(self, **kwargs):
        super(CircularProgressBar, self).__init__(**kwargs)
        Clock.schedule_once(self.animate, 0)

    def animate(self, *args):
        Clock.schedule_interval(self.percent_counter, self.duration / self.value)

    def percent_counter(self, *args):
        if self.counter < self.value:
            self.counter += 1
            self.text = f"{self.counter}%"
            self.set_value = self.counter
        else:
            Clock.unschedule(self.percent_counter)
            #ProgressBarScreen.reload_screen(self)


class SplashScreen(Screen):
    pass


ScreenManager = ScreenManager()
ScreenManager.add_widget(HomeScreen(name="Home"))
ScreenManager.add_widget(UploadScreen(name="Upload"))
ScreenManager.add_widget(TBDetectedScreen(name="Detected"))
ScreenManager.add_widget(TBNotDetectedScreen(name="Unverified"))
ScreenManager.add_widget(ProgressBarScreen(name="Progress"))
ScreenManager.add_widget(MultipleImagesScreen(name="Multiple"))
ScreenManager.add_widget(SplashScreen(name="Splash"))


class TBAPP(MDApp):

    global ScreenManager

    def build(self):
        self.theme_cls.theme_style_switch_animation = True
        self.theme_cls.theme_style_switch_animation_duration = 0.8
        self.theme_cls.theme_style = "Light"
        self.theme_cls.primary_palette = "Darkgray"

        #ScreenManager.add_widget(Builder.load_file('TBDetection.kv'))
        ScreenManager.add_widget(Builder.load_file("splash.kv"))

        return ScreenManager

    def switch_mode(self):
        self.theme_cls.primary_palette = (
            "Darkgray" if self.theme_cls.primary_palette == "Cyan" else "Cyan"
        )
        self.theme_cls.theme_style = (
            "Dark" if self.theme_cls.theme_style == "Light" else "Light"
        )

    def on_start(self):
        Clock.schedule_once(self.change_screen, 5)

    def change_screen(self, *args):
        ScreenManager.current = "Home"


if __name__ == '__main__':
    TBAPP().run()
